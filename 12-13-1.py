#/usr/bin/env python
""""""

import openpyxl, os
from openpyxl.styles import Font, NamedStyle

os.chdir("c:\\")
workbook = openpyxl.load_workbook("a.xlsx")
worksheet = workbook.get_sheet_by_name("Sheet1")
matrix = int(input("please input the colum-volum: "))
fontObj = Font(bold = True)
styleObj = NamedStyle(font = fontObj)

if matrix > 0:
    for row in range(1, matrix + 2):
        worksheet.row_dimensions[row].height = 15
        for colum in range(1, matrix + 2):
            if 1 == colum and 1 == row:
                pass
            elif 1 == row:
                print("aaa %d  %d" % (row,colum))
                # print(worksheet.cell(row=row, column=colum).column)
                worksheet.cell(row = row, column = colum).value = colum - 1
                worksheet.cell(row=row, column=colum).style = styleObj
            elif 1 == colum:
                print("bbb %d  %d" % (row, colum))
                worksheet.cell(row = row, column = colum).value = row - 1
                worksheet.cell(row=row, column=colum).style = styleObj
            else:
                # worksheet.cell(row=row, column=colum).value = (colum - 1) * (row - 1)
                cellObj = worksheet.cell(row=row, column=colum)
                print("=MMULT(%s:%s)" % (cellObj.column + str(1), "A" + str(row)))
                cellObj.value = "=MMULT(%s,%s)" % (cellObj.column + str(1), "A" + str(row))
workbook.save("d.xlsx")